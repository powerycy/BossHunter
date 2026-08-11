"""In-memory CSV/XLSX export for database-backed job selections."""

from __future__ import annotations

import csv
from copy import copy
import io
import json
import re
import sqlite3
from typing import Any, Iterable

from bosshunter.cities import get_city_code
from bosshunter.db import MAX_JOB_IDS, query_jobs


EXPORT_COLUMNS = [
	("岗位 ID", "id"),
	("职位", "title"),
	("公司", "company"),
	("薪资", "salary"),
	("城市", "city"),
	("BOSS 城市编码", "city_code"),
	("工作经验", "experience"),
	("JD", "jd"),
	("HR 姓名", "hr_name"),
	("HR 职位", "hr_title"),
	("HR 活跃度", "hr_active"),
	("公司规模", "company_size"),
	("公司行业", "company_industry"),
	("岗位链接", "url"),
	("预筛分", "quick_score"),
	("AI 分数", "score"),
	("评分理由", "score_reason"),
	("过滤来源", "filter_source"),
	("过滤原因", "filter_reason"),
	("人工推进", "manual_override"),
	("当前状态", "status"),
	("招呼语", "greeting"),
	("招呼语状态", "greeting_status"),
	("简历路径", "resume_path"),
	("创建时间", "created_at"),
	("更新时间", "updated_at"),
	("最近评分失败原因", "score_failure_reason"),
	("最近招呼语失败原因", "greeting_failure_reason"),
]


class InvalidJobSelectionError(ValueError):
	"""Raised when a selected export contains missing or deleted IDs."""

	def __init__(self, invalid_ids: list[str]):
		self.invalid_ids = invalid_ids
		super().__init__("所选岗位中存在无效、已删除或已过期的岗位 ID")


def _normalize_export_ids(job_ids: Iterable[Any] | None) -> list[str]:
	if job_ids is None:
		return []
	if isinstance(job_ids, (str, bytes, dict)):
		raise ValueError("岗位 ID 必须是数组")
	try:
		values = list(job_ids)
	except TypeError as exc:
		raise ValueError("岗位 ID 必须是数组") from exc
	ids: list[str] = []
	for value in values:
		if not isinstance(value, str):
			raise ValueError("岗位 ID 必须是字符串")
		job_id = value.strip()
		if job_id and job_id not in ids:
			ids.append(job_id)
	if len(ids) > MAX_JOB_IDS:
		raise ValueError(f"一次最多导出 {MAX_JOB_IDS} 个岗位")
	return ids


def _selected_rows(
	conn: sqlite3.Connection,
	scope: str,
	job_ids: Iterable[Any] | None,
	filters: dict[str, Any] | None,
) -> list[dict[str, Any]]:
	scope = str(scope or "all").strip().lower()
	if scope not in {"all", "filtered", "selected"}:
		raise ValueError("导出范围无效")
	ids = _normalize_export_ids(job_ids)
	if scope == "all":
		rows, _ = query_jobs(conn, deleted="active")
		return rows
	if scope == "filtered":
		rows, _ = query_jobs(conn, deleted="active", filters=filters or {})
		return rows
	if not ids:
		raise ValueError("所选岗位不能为空")
	rows, _ = query_jobs(conn, deleted="active", job_ids=ids)
	found = {str(row["id"]) for row in rows}
	invalid_ids = [job_id for job_id in ids if job_id not in found]
	if invalid_ids:
		raise InvalidJobSelectionError(invalid_ids)
	return rows


def _failure_reason(job: dict[str, Any]) -> str:
	value = job.get("score_failure_json")
	if value:
		try:
			payload = json.loads(str(value))
			if isinstance(payload, dict) and payload.get("message"):
				return str(payload["message"])[:240]
		except (TypeError, json.JSONDecodeError):
			pass
	reason = str(job.get("score_reason") or "")
	if reason.startswith(("AI评分失败:", "AI 评分失败:", "评分失败:")):
		return reason.split(":", 1)[1].strip()[:240]
	return ""


def _greeting_failure_reason(job: dict[str, Any]) -> str:
	value = job.get("greeting_failure_json")
	if value:
		try:
			payload = json.loads(str(value))
			if isinstance(payload, dict):
				return str(payload.get("user_message") or payload.get("message") or "")[:240]
		except (TypeError, json.JSONDecodeError):
			pass
	return ""


def _row_values(job: dict[str, Any]) -> list[Any]:
	city_code = str(job.get("city_code") or "").strip() or (get_city_code(str(job.get("city") or "")) or "")
	job = {**job, "city_code": city_code, "score_failure_reason": _failure_reason(job), "greeting_failure_reason": _greeting_failure_reason(job)}
	return [job.get(key, "") if job.get(key) is not None else "" for _, key in EXPORT_COLUMNS]


def _csv_safe(value: Any) -> str:
	text = "" if value is None else str(value)
	if text.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
		return "'" + text
	return text


def build_csv(rows: list[dict[str, Any]]) -> bytes:
	buffer = io.StringIO(newline="")
	writer = csv.writer(buffer, lineterminator="\r\n")
	writer.writerow([label for label, _ in EXPORT_COLUMNS])
	for job in rows:
		writer.writerow([_csv_safe(value) for value in _row_values(job)])
	return buffer.getvalue().encode("utf-8-sig")


def build_xlsx(rows: list[dict[str, Any]]) -> bytes:
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font
	except ImportError as exc:
		raise RuntimeError("导出 XLSX 需要安装 openpyxl") from exc
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "岗位池"
	headers = [label for label, _ in EXPORT_COLUMNS]
	sheet.append(headers)
	for cell in sheet[1]:
		cell.font = Font(bold=True)
	for job in rows:
		values = _row_values(job)
		sheet.append(values)
		url_column = next(index for index, (_, key) in enumerate(EXPORT_COLUMNS, start=1) if key == "url")
		cell = sheet.cell(sheet.max_row, url_column)
		if cell.value:
			cell.hyperlink = str(cell.value)
			cell.style = "Hyperlink"
	for index, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
		max_length = max(len(str(sheet.cell(row, index).value or "")) for row in range(1, min(sheet.max_row, 30) + 1))
		sheet.column_dimensions[chr(64 + index) if index <= 26 else sheet.cell(1, index).column_letter].width = min(max(max_length + 2, 12), 48)
		if key in {"jd", "score_reason", "greeting"}:
			for row in range(2, sheet.max_row + 1):
				cell = sheet.cell(row, index)
				alignment = copy(cell.alignment)
				alignment.wrap_text = True
				cell.alignment = alignment
	sheet.freeze_panes = "A2"
	sheet.auto_filter.ref = sheet.dimensions
	output = io.BytesIO()
	workbook.save(output)
	return output.getvalue()


def export_jobs(
	conn: sqlite3.Connection,
	*,
	format: str = "xlsx",
	scope: str = "all",
	job_ids: Iterable[Any] | None = None,
	filters: dict[str, Any] | None = None,
) -> tuple[bytes, str, str]:
	format = str(format or "xlsx").strip().lower()
	if format not in {"csv", "xlsx"}:
		raise ValueError("导出格式只支持 xlsx 或 csv")
	rows = _selected_rows(conn, scope, job_ids, filters)
	if format == "csv":
		return build_csv(rows), "text/csv; charset=utf-8", "bosshunter-jobs.csv"
	if format == "xlsx":
		return build_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "bosshunter-jobs.xlsx"


def export_row_count(
	conn: sqlite3.Connection,
	*,
	scope: str = "all",
	job_ids: Iterable[Any] | None = None,
	filters: dict[str, Any] | None = None,
) -> int:
	return len(_selected_rows(conn, scope, job_ids, filters))
