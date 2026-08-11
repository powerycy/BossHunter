import io
import importlib.util
import json
import csv
import sqlite3
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from bosshunter.ai.credentials import normalize_openai_base_url
from bosshunter.ai.diagnostics import diagnose_ai
from bosshunter.cities import CityRefreshError, load_city_snapshot, refresh_city_cache
from bosshunter.db import (
    JobDeletionConflictError,
    JobDeletionConfirmationError,
    add_history,
    get_db,
    get_jobs_by_status,
    get_jobs_pending_confirmation,
    get_jobs_ready_to_send,
    insert_job,
    job_exists,
    job_url_exists,
    permanent_delete_jobs,
    query_jobs,
    restore_jobs,
    soft_delete_jobs,
    update_job_greeting,
    update_job_score,
    update_job_score_failure,
    update_job_status,
)
from bosshunter.job_export import build_csv, build_xlsx, export_jobs
from bosshunter.scoring_selection import preview_scoring, select_scoring_jobs


HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


def _job(job_id: str, city: str = "北京", company: str = "Example") -> dict:
    return {
        "id": job_id,
        "title": "Engineer",
        "company": company,
        "salary": "10-20K",
        "city": city,
        "experience": "1-3 years",
        "jd": "Build product features",
        "hr_name": "HR",
        "hr_title": "Recruiter",
        "hr_active": "active",
        "company_size": "100-499",
        "company_industry": "Software",
        "url": f"https://example.com/jobs/{job_id}",
    }


class CitySnapshotTests(unittest.TestCase):
    def test_bundled_snapshot_is_complete_and_offline(self):
        with tempfile.TemporaryDirectory() as tmp:
            snapshot = load_city_snapshot(cache_path=Path(tmp) / "cities.cache.json")

        self.assertEqual(snapshot["source"], "bundled")
        self.assertGreaterEqual(len(snapshot["cities"]), 300)
        codes = {city["name"]: city["code"] for city in snapshot["cities"]}
        self.assertEqual(codes["\u5317\u4eac"], "101010100")
        self.assertEqual(codes["\u4e0a\u6d77"], "101020100")
        self.assertEqual(codes["\u5e7f\u5dde"], "101280100")
        self.assertEqual(codes["\u6df1\u5733"], "101280600")

    def test_invalid_cache_falls_back_without_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            cache_path = Path(tmp) / "cities.cache.json"
            cache_path.write_text("<html>not-json</html>", encoding="utf-8")
            with patch("bosshunter.cities.httpx.get") as http_get:
                snapshot = load_city_snapshot(cache_path=cache_path)

        self.assertEqual(snapshot["source"], "bundled")
        http_get.assert_not_called()

    def test_html_refresh_is_rejected_without_overwriting_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            cache_path = Path(tmp) / "cities.cache.json"
            original = load_city_snapshot(cache_path=cache_path)
            cache_path.write_text(json.dumps(original, ensure_ascii=False), encoding="utf-8")
            response = SimpleNamespace(
                status_code=200,
                headers={"content-type": "text/html"},
                text="<html>blocked</html>",
            )

            with self.assertRaises(CityRefreshError):
                refresh_city_cache(cache_path, fetcher=lambda *args, **kwargs: response)

            self.assertEqual(json.loads(cache_path.read_text(encoding="utf-8"))["schema"], "bosshunter.cities.v1")


class AiDiagnosticsTests(unittest.TestCase):
    def test_lulucoding_url_normalization_never_duplicates_v1(self):
        for value in (
            "https://api.lulucoding.com",
            "https://api.lulucoding.com/v1",
            "https://api.lulucoding.com/v1/v1",
            "https://api.lulucoding.com/v1/chat/completions",
        ):
            self.assertEqual(
                normalize_openai_base_url(value, "lulucoding"),
                "https://api.lulucoding.com/v1",
            )

    @patch("bosshunter.ai.diagnostics.httpx.get")
    def test_basic_diagnostics_only_reads_models_and_never_calls_chat(self, http_get):
        response = SimpleNamespace(
            status_code=200,
            headers={"content-type": "application/json"},
            text='{"data":[{"id":"lulu-model"}]}',
            json=lambda: {"data": [{"id": "lulu-model"}]},
        )
        http_get.return_value = response
        config = {
            "ai": {
                "service": "lulucoding",
                "provider": "openai_compatible",
                "base_url": "https://api.lulucoding.com/v1/v1",
                "api_key": "local-test-key",
                "model": "lulu-model",
            }
        }

        with patch("bosshunter.ai.diagnostics.httpx.post") as http_post:
            result = diagnose_ai(config)

        self.assertTrue(result["ok"])
        self.assertTrue(result["billable"] is False)
        self.assertEqual(http_get.call_args.args[0], "https://api.lulucoding.com/v1/models")
        http_post.assert_not_called()

    @patch("bosshunter.ai.diagnostics.httpx.get")
    def test_html_models_response_is_classified_without_echoing_key(self, http_get):
        http_get.return_value = SimpleNamespace(
            status_code=200,
            headers={"content-type": "text/html"},
            text="<html>provider error</html>",
        )
        result = diagnose_ai({
            "ai": {
                "service": "lulucoding",
                "provider": "openai_compatible",
                "api_key": "local-secret-key",
                "model": "lulu-model",
                "base_url": "https://api.lulucoding.com",
            }
        })

        self.assertFalse(result["ok"])
        self.assertEqual(result["error_kind"], "invalid_content_type")
        self.assertNotIn("local-secret-key", json.dumps(result))


class ScoringSelectionTests(unittest.TestCase):
    def test_scopes_skip_delivery_locked_and_identify_retryable_failures(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "jobs.db")
            try:
                for job_id in ("pending", "failed", "ready", "sent"):
                    insert_job(db, _job(job_id))
                update_job_score(db, "failed", 0, "AI score failed: temporary")
                update_job_score_failure(db, "failed", json.dumps({"message": "temporary", "retryable": True}))
                update_job_status(db, "ready", "ready")
                update_job_score(db, "ready", 88, "good match")
                update_job_status(db, "sent", "sent")

                pending = select_scoring_jobs(db, scope="pending", limit=None)
                failed = select_scoring_jobs(db, scope="failed", limit=None)
                scored = select_scoring_jobs(db, scope="all_scored", limit=None)
                preview = preview_scoring(db, scope="pending", limit=20, max_attempts_per_job=2)
            finally:
                db.close()

        self.assertEqual({job["id"] for job in pending}, {"pending", "failed"})
        self.assertEqual([job["id"] for job in failed], ["failed"])
        self.assertEqual([job["id"] for job in scored], ["ready"])
        self.assertEqual(preview["first_attempt_requests"], 2)
        self.assertEqual(preview["max_possible_requests"], 4)


class JobExportTests(unittest.TestCase):
    def test_all_filtered_and_selected_exports_have_exact_database_id_sets(self):
        if not HAS_OPENPYXL:
            self.skipTest("openpyxl dependency is not installed in this interpreter")

        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "export-ranges.db")
            try:
                for index in range(8):
                    job = _job(f"range-{index}", city="北京" if index < 4 else "上海")
                    job["title"] = "Python 后端" if index in {0, 1, 2} else "Java 后端"
                    job["status"] = "ready" if index in {0, 1, 2} else "pending"
                    insert_job(db, job)
                    if job["status"] != "pending":
                        update_job_status(db, job["id"], job["status"])

                for export_format in ("csv", "xlsx"):
                    all_content, _, _ = export_jobs(db, format=export_format, scope="all")
                    filtered_content, _, _ = export_jobs(
                        db,
                        format=export_format,
                        scope="filtered",
                        job_ids=["range-0"],
                        filters={"q": "Python", "city": "北京", "status": "ready"},
                    )
                    selected_content, _, _ = export_jobs(
                        db,
                        format=export_format,
                        scope="selected",
                        job_ids=["range-1", "range-7"],
                    )
                    exported_sets = {
                        "all": self._exported_ids(export_format, all_content),
                        "filtered": self._exported_ids(export_format, filtered_content),
                        "selected": self._exported_ids(export_format, selected_content),
                    }
                    self.assertEqual(exported_sets["all"], {f"range-{i}" for i in range(8)})
                    self.assertEqual(exported_sets["filtered"], {"range-0", "range-1", "range-2"})
                    self.assertEqual(exported_sets["selected"], {"range-1", "range-7"})
            finally:
                db.close()

    @staticmethod
    def _exported_ids(export_format: str, content: bytes) -> set[str]:
        if export_format == "csv":
            rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
            id_column = rows[0].index("岗位 ID")
            return {row[id_column] for row in rows[1:]}
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=False)
        try:
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            id_column = headers.index("岗位 ID") + 1
            return {str(sheet.cell(row, id_column).value) for row in range(2, sheet.max_row + 1)}
        finally:
            workbook.close()

    def test_filtered_export_queries_structured_filters_not_current_page_ids(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "export.db")
            try:
                for index in range(8):
                    job = _job(f"filter-{index}", city="北京" if index < 4 else "上海")
                    job["title"] = "Python 后端" if index in {0, 1, 2} else "Java 后端"
                    job["status"] = "ready" if index in {0, 1, 2} else "pending"
                    insert_job(db, job)
                    if job["status"] != "pending":
                        update_job_status(db, job["id"], job["status"])

                content, _, _ = export_jobs(
                    db,
                    format="csv",
                    scope="filtered",
                    job_ids=["filter-0"],  # 模拟当前分页只把一条 ID 传回前端
                    filters={"q": "Python", "city": "北京", "status": "ready"},
                )
            finally:
                db.close()

        csv_rows = list(csv.reader(io.StringIO(content.decode("utf-8-sig"))))
        id_column = csv_rows[0].index("岗位 ID")
        self.assertEqual({row[id_column] for row in csv_rows[1:]}, {"filter-0", "filter-1", "filter-2"})

    def test_csv_has_bom_formula_protection_and_resolved_city_code(self):
        job = _job("csv-1", company="=2+2")
        job["salary"] = "-1"
        csv_bytes = build_csv([job])

        self.assertTrue(csv_bytes.startswith(b"\xef\xbb\xbf"))
        csv_text = csv_bytes.decode("utf-8-sig")
        self.assertIn("'=2+2", csv_text)
        self.assertIn("'-1", csv_text)
        self.assertIn("101010100", csv_text)
        self.assertIn("https://example.com/jobs/csv-1", csv_text)

    @unittest.skipUnless(HAS_OPENPYXL, "openpyxl dependency is not installed in this interpreter")
    def test_xlsx_contains_clickable_job_url(self):
        job = _job("xlsx-1", city="\u4e0a\u6d77")
        workbook_bytes = build_xlsx([job])

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=False)
        try:
            sheet = workbook.active
            url_column = next(cell.column for cell in sheet[1] if cell.value == "\u5c97\u4f4d\u94fe\u63a5")
            url_cell = sheet.cell(2, url_column)
            self.assertEqual(url_cell.value, "https://example.com/jobs/xlsx-1")
            self.assertEqual(url_cell.hyperlink.target, "https://example.com/jobs/xlsx-1")
        finally:
            workbook.close()


class JobDeletionDataTests(unittest.TestCase):
    def test_legacy_database_migrates_idempotently_without_rewriting_job_data(self):
        with tempfile.TemporaryDirectory() as tmp:
            db_path = Path(tmp) / "legacy.db"
            legacy = sqlite3.connect(db_path)
            legacy.executescript(
                """
                CREATE TABLE jobs (
                    id TEXT PRIMARY KEY, title TEXT NOT NULL, company TEXT NOT NULL,
                    salary TEXT, city TEXT, experience TEXT, jd TEXT, hr_name TEXT,
                    hr_title TEXT, hr_active TEXT, company_size TEXT, company_industry TEXT,
                    url TEXT, score INTEGER DEFAULT 0, score_reason TEXT, greeting TEXT,
                    status TEXT DEFAULT 'pending', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, job_id TEXT NOT NULL,
                    action TEXT NOT NULL, detail TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                INSERT INTO jobs (id, title, company, city, url, score, greeting, status)
                VALUES ('legacy', 'Legacy Job', 'Legacy Co', 'Beijing', 'https://example.com/legacy', 77, 'hello', 'ready');
                INSERT INTO history (job_id, action, detail) VALUES ('legacy', 'approved', 'kept');
                """
            )
            legacy.commit()
            legacy.close()

            first = get_db(db_path)
            try:
                columns = {row[1] for row in first.execute("PRAGMA table_info(jobs)").fetchall()}
                row = first.execute("SELECT id, title, score, greeting, status, deleted_at, deleted_reason FROM jobs WHERE id='legacy'").fetchone()
                self.assertTrue({"deleted_at", "deleted_reason"}.issubset(columns))
                self.assertEqual(dict(row)["title"], "Legacy Job")
                self.assertEqual(dict(row)["score"], 77)
                self.assertEqual(dict(row)["greeting"], "hello")
                self.assertIsNone(dict(row)["deleted_at"])
            finally:
                first.close()

            second = get_db(db_path)
            try:
                columns_again = [row[1] for row in second.execute("PRAGMA table_info(jobs)").fetchall()]
                self.assertEqual(columns_again.count("deleted_at"), 1)
                self.assertEqual(second.execute("SELECT COUNT(*) FROM jobs").fetchone()[0], 1)
                self.assertEqual(second.execute("SELECT COUNT(*) FROM history").fetchone()[0], 1)
            finally:
                second.close()

    def test_soft_delete_restore_preserves_business_fields_and_excludes_queues(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "delete.db")
            try:
                job = _job("delete-preserve")
                insert_job(db, job)
                update_job_score(db, job["id"], 88, "保留评分理由")
                update_job_greeting(db, job["id"], "保留招呼语")
                update_job_status(db, job["id"], "approved")
                add_history(db, job["id"], "approved", "保留历史")
                before = dict(db.execute("SELECT * FROM jobs WHERE id = ?", (job["id"],)).fetchone())

                self.assertFalse(soft_delete_jobs(db, [job["id"]], confirmed=False))
            except JobDeletionConfirmationError:
                pass

            result = soft_delete_jobs(db, [job["id"]], confirmed=True, reason="测试移入回收站")
            self.assertEqual(result["affected_count"], 1)
            deleted = dict(db.execute("SELECT * FROM jobs WHERE id = ?", (job["id"],)).fetchone())
            self.assertEqual(deleted["status"], before["status"])
            self.assertEqual(deleted["score"], before["score"])
            self.assertEqual(deleted["greeting"], before["greeting"])
            self.assertEqual(deleted["url"], before["url"])
            self.assertIsNotNone(deleted["deleted_at"])
            self.assertEqual(query_jobs(db, deleted="active")[1], 0)
            self.assertEqual(query_jobs(db, deleted="only")[1], 1)
            self.assertEqual(get_jobs_by_status(db, "approved"), [])
            self.assertEqual(get_jobs_pending_confirmation(db), [])
            self.assertEqual(get_jobs_ready_to_send(db), [])
            self.assertEqual(job_exists(db, job["id"]), True)
            self.assertEqual(job_url_exists(db, job["url"]), True)
            self.assertEqual(insert_job(db, job), False)
            self.assertEqual([row["id"] for row in db.execute("SELECT * FROM jobs WHERE deleted_at IS NULL").fetchall()], [])

            restored = restore_jobs(db, [job["id"]], confirmed=True)
            self.assertEqual(restored["affected_count"], 1)
            restored_again = restore_jobs(db, [job["id"]], confirmed=True)
            self.assertEqual(restored_again["affected_count"], 0)
            after = dict(db.execute("SELECT * FROM jobs WHERE id = ?", (job["id"],)).fetchone())
            self.assertEqual(after["status"], before["status"])
            self.assertEqual(after["score"], before["score"])
            self.assertEqual(after["greeting"], before["greeting"])
            self.assertIsNone(after["deleted_at"])
            actions = {row["action"] for row in db.execute("SELECT action FROM history WHERE job_id = ?", (job["id"],)).fetchall()}
            self.assertTrue({"soft_deleted", "restored"}.issubset(actions))
            db.close()

    def test_permanent_delete_protects_status_history_and_task_checkpoint_atomically(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "delete-protect.db")
            try:
                protected = _job("protected")
                historical = _job("historical")
                safe = _job("safe")
                checkpointed = _job("checkpointed")
                for job in (protected, historical, safe, checkpointed):
                    insert_job(db, job)
                update_job_status(db, "protected", "sent")
                add_history(db, "protected", "sent", "发送证据")
                add_history(db, "historical", "sent", "历史发送证据")
                update_job_status(db, "historical", "pending")
                for job_id in ("protected", "historical", "safe", "checkpointed"):
                    soft_delete_jobs(db, [job_id], confirmed=True)
                db.execute(
                    """INSERT INTO workbench_tasks
                       (id, mode, label, status, config_snapshot_json, checkpoint_json, progress_json, logs_json, context_refs_json)
                       VALUES (?, 'score', '测试任务', 'paused', '{}', ?, '{}', '[]', '{}')""",
                    ("checkpoint-task", json.dumps({"remaining_job_ids": ["checkpointed"]})),
                )
                db.commit()

                with self.assertRaises(JobDeletionConflictError) as status_error:
                    permanent_delete_jobs(db, ["protected"], confirmed=True, confirmation="PERMANENT_DELETE")
                self.assertTrue(status_error.exception.blocked)
                with self.assertRaises(JobDeletionConflictError) as history_error:
                    permanent_delete_jobs(db, ["historical"], confirmed=True, confirmation="PERMANENT_DELETE")
                self.assertTrue(history_error.exception.blocked)
                with self.assertRaises(JobDeletionConflictError) as task_error:
                    permanent_delete_jobs(db, ["checkpointed"], confirmed=True, confirmation="PERMANENT_DELETE")
                self.assertTrue(task_error.exception.blocked)

                with self.assertRaises(JobDeletionConflictError):
                    permanent_delete_jobs(db, ["safe", "protected"], confirmed=True, confirmation="PERMANENT_DELETE")
                self.assertIsNotNone(db.execute("SELECT 1 FROM jobs WHERE id = 'safe'").fetchone())

                deleted = permanent_delete_jobs(db, ["safe"], confirmed=True, confirmation="PERMANENT_DELETE")
                self.assertEqual(deleted["affected_count"], 1)
                self.assertIsNone(db.execute("SELECT 1 FROM jobs WHERE id = 'safe'").fetchone())
                self.assertIsNone(db.execute("SELECT 1 FROM history WHERE job_id = 'safe'").fetchone())
            finally:
                db.close()

    def test_permanent_delete_requires_strong_confirmation_without_mutation(self):
        with tempfile.TemporaryDirectory() as tmp:
            db = get_db(Path(tmp) / "delete-confirm.db")
            try:
                job = _job("confirm")
                insert_job(db, job)
                soft_delete_jobs(db, [job["id"]], confirmed=True)
                with self.assertRaises(JobDeletionConfirmationError):
                    permanent_delete_jobs(db, [job["id"]], confirmed=True, confirmation="DELETE")
                self.assertIsNotNone(db.execute("SELECT 1 FROM jobs WHERE id = 'confirm'").fetchone())
            finally:
                db.close()


if __name__ == "__main__":
    unittest.main()
